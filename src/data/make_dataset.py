import os
import pandas as pd
from openpyxl import load_workbook
from .filtering_function import convert_object_columns_to_integers

def importation_excel(excel_file_path, sheet_name):
    """
    Imports data from a specified Excel file.

    Args:
        excel_file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet in the Excel file.

    Returns:
        pd.DataFrame: Pandas DataFrame containing the data from the specified sheet.
    
    Raises:
        AssertionError: If the Excel file or the specified sheet does not exist.
        AssertionError: If column names are empty.

    Example:
        >>> import_excel_data('my_file.xlsx', 'Sheet1')
    """

    # Check if the Excel file exists
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"The Excel file '{excel_file_path}' does not exist.")

    # Load the Excel workbook in read-only mode
    workbook = load_workbook(excel_file_path, read_only=True, data_only=True)

    # Check if the specified sheet exists in the Excel file
    if not sheet_name in workbook.sheetnames:
        raise ValueError(f"The sheet '{sheet_name}' does not exist in the Excel file.")

    # Select the specified sheet
    sheet = workbook[sheet_name]

    # Get column names from the first row of the sheet
    column_names = [cell.value for cell in sheet[1]]
    # Check if column names are not empty
    if not all(name is not None for name in column_names):
        raise ValueError("Column names cannot be empty.")

    # Get data from the second row of the sheet
    data_row = [cell.value for cell in sheet[2]]

    # Create a Pandas DataFrame with the data and column names
    df = pd.DataFrame([data_row], columns=column_names)

    df = convert_object_columns_to_integers(df)
    # Return the created DataFrame
    return df